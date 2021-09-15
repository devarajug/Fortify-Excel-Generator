import pandas as pd
from sys import argv
from os.path import splitext
from openpyxl import Workbook
import xml.etree.ElementTree as ET
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill

class XmlToXls:

    def __init__(self, in_file, out_file):
        self.category = []
        self.priority = []
        self.kingdom = []
        self.abstract = []
        self.source_file_path = []
        self.source_file_num = []
        self.sink_file_path = []
        self.sink_file_no = []
        self.status = []
        self.auditor_comment = []
        self.in_file = in_file
        self.outfile = out_file
    
    def zipXmlData(self):
        tree = ET.parse(self.file)
        root = tree.getroot()
        for issue in root.findall("ReportSection/SubSection/IssueListing/Chart/GroupingSection/Issue"):
            auditor_comment_value = ""
            for i in range(len(issue)):
                if issue[i].tag == "Category":
                    category_value = issue[i].text
                elif issue[i].tag == "Kingdom":
                    kingdom_value = issue[i].text
                elif issue[i].tag == "Friority":
                    priority_value = issue[i].text
                elif issue[i].tag == "Abstract":
                    abstract_value = issue[i].text
                elif issue[i].tag == "Tag":
                    status_value = issue[i][1].text
                elif issue[i].tag == "Comment":
                    auditor_comment_value+= str(issue[i][0].text) +"\n"+ str(issue[i][1].text)
                elif issue[i].tag == "Source":
                    source_file_path_value = issue[i][1].text
                    source_line_no_value = issue[i][2].text
                elif issue[i].tag == "Primary":
                    sink_file_path_value = issue[i][1].text
                    sink_line_no_value = issue[i][2].text

            self.category.append(category_value)
            self.kingdom.append(kingdom_value)
            self.priority.append(priority_value)
            self.abstract.append(abstract_value)
            self.source_file_path.append(source_file_path_value)
            self.source_file_num.append(source_line_no_value)
            self.sink_file_path.append(sink_file_path_value)
            self.sink_file_no.append(sink_line_no_value)
            self.status.append(status_value)
            self.auditor_comment.append(auditor_comment_value)
        
        return zip(
            self.category,
            self.priority,
            self.kingdom,
            self.abstract,
            self.source_file_path,
            self.source_file_num,
            self.sink_file_path,
            self.sink_file_no,
            self.status,
            self.auditor_comment
        )
    
    def makeXL(self):
        zipedxmldata = list(self.zipXmlData())
        workbook = Workbook()
        workbook.remove(workbook.active)
        header_font = Font(name='Calibri',bold=True,color='FFFFFF')
        centered_alignment = Alignment(horizontal='center')
        wrapped_alignment = Alignment(vertical='top',wrap_text=False)
        sast_sheet_columns = [
            ("Category", 40),
            ("Priority", 15),
            ("Kingdom", 35),
            ("Abstract", 35),
            ("Source File Path", 30),
            ("Source File No", 15),
            ("Sink File Path", 30),
            ("Sink File No", 15),
            ("Status", 20),
            ("Auditor Comment", 40),
            ("Developer Comment", 40),
        ]

        worksheet = workbook.create_sheet(title='Vulnerability',index=0)
        fill = PatternFill(start_color='5FABE6',end_color='5FABE6',fill_type='solid',)
        row_num = 1
        for col_num, (column_title, column_width) in enumerate(sast_sheet_columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.font = header_font
            cell.alignment = centered_alignment
            cell.fill = fill
            column_letter = get_column_letter(col_num)
            column_dimensions = worksheet.column_dimensions[column_letter]
            column_dimensions.width = column_width
        for row in zipedxmldata:
            row_num+=1
            row = [
                (row[0], 'Normal'),
                (row[1], 'Normal'),
                (row[2], 'Normal'),
                (row[3], 'Normal'),
                (row[4], 'Normal'),
                (row[5], 'Normal'),
                (row[6], 'Normal'),
                (row[7], 'Normal'),
                (row[8], 'Normal'),
                (row[9], 'Normal'),
            ]

            for col_num, (cell_value, cell_format) in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
                cell.style = cell_format
                cell.alignment = wrapped_alignment
        
        worksheet.freeze_panes = worksheet['A2']
        worksheet.sheet_properties.tabColor = '5FABE6'
        workbook.save(self.out_file_name)
        self.makeHtml()
        print('execl created successfully....')
        return ""
    
    def makeHtml(self):
        xl = pd.ExcelFile(self.out_file_name)
        dfxl = xl.parse(xl.sheet_names[0])
        dfopen = dfxl[(dfxl['Status'] == 'Open') | (dfxl['Status'] == 'open')]
        dfnotanissue = dfxl[(dfxl['Status'] == 'Not an Issue') | (dfxl['Status'] == 'Not an issue')]
        dfclosed = dfxl[(dfxl['Status'] == 'Closed') | (dfxl['Status'] == 'closed')]

        OpenCritical        = len(dfopen[dfopen['Priority'] == "Critical"])
        OpenHigh            = len(dfopen[dfopen['Priority'] == "High"])
        OpenMedium          = len(dfopen[dfopen['Priority'] == "Medium"])
        OpenLow             = len(dfopen[dfopen['Priority'] == "Low"])

        NotAnIssueCritical  = len(dfnotanissue[dfnotanissue['Priority'] == "Critical"])
        NotAnIssueHigh      = len(dfnotanissue[dfnotanissue['Priority'] == "High"])
        NotAnIssueMedium    = len(dfnotanissue[dfnotanissue['Priority'] == "Medium"])
        NotAnIssueLow       = len(dfnotanissue[dfnotanissue['Priority'] == "Low"])

        ClosedCritical      = len(dfclosed[dfclosed['Priority'] == "Critical"])
        ClosedHigh          = len(dfclosed[dfclosed['Priority'] == "High"])
        ClosedMedium        = len(dfclosed[dfclosed['Priority'] == "Medium"])
        ClosedLow           = len(dfclosed[dfclosed['Priority'] == "Low"])

        CriticalTotal       = OpenCritical+ClosedCritical+NotAnIssueCritical
        HighTotal           = OpenHigh+ClosedHigh+NotAnIssueHigh
        MediumTotal         = OpenMedium+ClosedMedium+NotAnIssueMedium
        LowTotal            = OpenLow+ClosedLow+NotAnIssueLow
        GrandTotal          = len(dfclosed)+len(dfnotanissue)+len(dfopen)
        print("Open :", len(dfopen))
        print("Not an Issue :", len(dfnotanissue))
        print("Closed :", len(dfclosed))
        print()
        html_template = """<style>
        .issuestable {{font-family: Arial, Helvetica, sans-serif; border-collapse: collapse; width: 50%; }}
        .issuestable td, .customers th {{border: 1px solid #ddd; padding: 8px; }}
        .issuestable tr:nth-child(even){{background-color: #f2f2f2;}}
        .issuestable tr:hover {{background-color: #ddd;}}
        .issuestable th {{padding-top: 12px; padding-bottom: 12px; text-align: center; background-color: #02bdc7; color: white; }}
        </style>
        <table class="issuestable">
            <tr><th>Status</th><th>Critical</th><th>High</th><th>Medium</th><th>Low</th><th>Total</th></tr>
            <tr><td>Closed</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td></tr>
            <tr><td>Not an Issue</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td></tr>
            <tr><td>Open</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td></tr>
            <tr><td>Grand Total</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td></tr>
        </table>""".format(
            ClosedCritical, ClosedHigh, ClosedMedium, ClosedLow, len(dfclosed),
            NotAnIssueCritical, NotAnIssueHigh, NotAnIssueMedium, NotAnIssueLow, len(dfnotanissue),
            OpenCritical, OpenHigh, OpenMedium, OpenLow, len(dfopen),
            CriticalTotal, HighTotal, MediumTotal, LowTotal, GrandTotal
        )

        with open(str(splitext(self.out_file_name)[0])+".html", 'w') as wr:
            wr.write(html_template)
        table_data = [
            ("Status", "Critical", "High", "Medium", "Low", "Total"),
            ("Closed", ClosedCritical, ClosedHigh, ClosedMedium, ClosedLow, len(dfclosed)),
            ("Not an Issue", NotAnIssueCritical, NotAnIssueHigh, NotAnIssueMedium, NotAnIssueLow, len(dfnotanissue)),
            ("Open", OpenCritical, OpenHigh, OpenMedium, OpenLow, len(dfopen)),
            ("Grand Total", CriticalTotal, HighTotal, MediumTotal, LowTotal, GrandTotal)
        ]
        print("+---------------+----------+----------+----------+----------+----------+")
        for row in table_data:
            print("|{: ^15}|{: ^10}|{: ^10}|{: ^10}|{: ^10}|{: ^10}|".format(*row))
            print("+---------------+----------+----------+----------+----------+----------+")
        print()

if __name__ == "__main__":
    file = argv[1]
    outfile_name = str(splitext(file)[0])+".xlsx"
    XmlToXls(file=file, out_file_name=outfile_name).makeXL()



